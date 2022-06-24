from neo4j import GraphDatabase
import openpyxl
from sympy import continued_fraction
from datetime import datetime

class Neo4jConnection:
    
    def __init__(self, uri, user, pwd):
        self.__uri = uri
        self.__user = user
        self.__pwd = pwd
        self.__driver = None
        try:
            self.__driver = GraphDatabase.driver(self.__uri, auth=(self.__user, self.__pwd))
            print("Driver created")
        except Exception as e:
            print("Failed to create the driver:", e)
        
    def close(self):
        if self.__driver is not None:
            self.__driver.close()
        
    def query(self, query, db=None):
        assert self.__driver is not None, "Driver not initialized!"
        session = None
        response = None
        try: 
            session = self.__driver.session(database=db) if db is not None else self.__driver.session() 
            response = list(session.run(query))
        except Exception as e:
            print("Query failed:", e)
        finally: 
            if session is not None:
                session.close()
        return response

if __name__ == "__main__":
    #connection with the database
    conn = Neo4jConnection(uri="bolt://localhost:7687", user="neo4j", pwd="ernest")

    #query_str = "CREATE (: Team { Name: 'Indipendente', Country: 'Argentina', Trophies: '21'} ) "

    #file to load
    path_to_file = input("Enter the file's path to load: ")  
    print(path_to_file)
    wookbook = openpyxl.load_workbook(path_to_file)

    worksheet = wookbook.active

    # option whether to add edges or nodes
    wtd_option= input("Enter N for node, E for edge, CN correct node, CE for correct edge:")

    if wtd_option == "N":
        print("Adding node")

    #   Name of the node;
    #   Name(as attribute); other attributes + values for each attribute
    #   attributes = []
    #   values = []         where value[i] denotes the value of the tuple for the attribute[i]

        attributes = []
        values = []
        last_read_n = 0 # for resuming when the file is not readed completely
        # Iterate the loop to read the cell values
        for i in range(0, worksheet.max_row): # loop over the rows
            if (i == 0): # CASE when we have the first row (update attributes)
                for col in worksheet.iter_cols(0, worksheet.max_column):
                    attributes.append(col[i].value)
            else: # CASE for rows other than first (update values)
                if i < last_read_n or i>100:
                    continue
                for col in worksheet.iter_cols(0, worksheet.max_column):
                    values.append(col[i].value)
                print(attributes)
                print(values)
                if type(values[0]) == str:
                    values[0]=values[0].replace(" ","")
                # have to create the part1 string for the query    
                p1 = "CREATE( :"+ attributes[0]

                # have to create the part2 string for the query(attributes)
                # {"+ attributes[0] +": '"+ values[0] + "' ,"+  attributes[1] +": '" + values[1] + "' ,"+ attributes[2] +":"+ str(values[2]) +"}
                # {Name: 'Indipendente' , Country: 'Argentina'}
                p2 = "{"
                first_attribute = True
                for i in range(0,len(values)):
                    if type(values[i]) != 'str':
                        values[i] = str(values[i])
                    values[i]=values[i].replace("'","")
                    str_concat = attributes[i] +": '"+ values[i] +"'"
                    if first_attribute == True:
                        first_attribute = False
                        p2 = p2 + str_concat
                    else:
                        p2 = p2 +"," + str_concat
                p2 = p2 +"} )"
                values = []
                conn.query(p1+p2,  db='prova')

    elif wtd_option == "E":
        print("Adding edge")
        # (Keanu)-[:ACTED_IN {roles:['Neo']}]->(TheMatrix)
        # match (p:Person) where p.Age=22 match (t:Name) where  create (p) - [:Likes] -> (t)
        relation_name = []
        attributes = []
        values = []
        last_read_e = 0  # for resuming when the file is not readed completely
        # Iterate the loop to read the cell values
        for i in range(0, worksheet.max_row): # loop over the rows
            print(str(int(i* 100 /1041850)) + "%")
            if (i == 0): # CASE relation name
                for col in worksheet.iter_cols(0, worksheet.max_column):
                    relation_name.append(col[i].value)
            elif (i == 1): # CASE when we have the first row (update attributes)
                for col in worksheet.iter_cols(0, worksheet.max_column):
                    attributes.append(col[i].value)
            else: # CASE for rows other than first (update values)
                if i < last_read_e:
                    continue
                for col in worksheet.iter_cols(0, worksheet.max_column):
                    values.append(col[i].value)
                print(relation_name)
                print(attributes)
                print(values)
                # create the outgoing part
                p1 = "MATCH (first:"+attributes[0] +") where first."+ attributes[0] +"= '"+ str(values[0]) +"' "
                p2 = ""
                relation_att =""
                # create the link part
                first_attribute = True
                for j in range(1,len(values)):
                    print(type(values[j]))
                    if type(values[j]) != str:
                        values[j] = str(values[j])
                    if (j == 1):
                        # create the ingoing node
                        p2 = " MATCH (second:"+attributes[j] +") where second."+ attributes[j] +"= '"+ values[j] +"'"
                    else:
                        if first_attribute == True:
                            first_attribute = False
                            relation_att = relation_att + attributes[j] +":"+ values[j]
                        else:
                            relation_att = relation_att +","+ attributes[j] +":"+ values[j]
                p3 = " create (first) - [:"+ relation_name[1] + "{"+ relation_att +"} ] -> (second)"
                conn.query(p1+p2+p3,  db='footballdb')
                values = []


    elif wtd_option == "CN":
        print("Correct nodes")
        #result=conn.query('''MATCH (n:ClubID) WHERE n.ClubID="1039" RETURN count(n)''',  db='prova')
        #print(result[0])
        #for record in result :
            #if(record["count(n)"] == 1):
                #print("ok")
            #else:
                #print("ko")

        total = 1
        attributes = []
        values = []
        last_read_n = 0 # for resuming when the file is not readed completely
        # Iterate the loop to read the cell values
        for i in range(0, worksheet.max_row): # loop over the rows
            if (i == 0): # CASE when we have the first row (update attributes)
                for col in worksheet.iter_cols(0, worksheet.max_column):
                    attributes.append(col[i].value)
            else: # CASE for rows other than first (update values)
                for col in worksheet.iter_cols(0, worksheet.max_column):
                    values.append(col[i].value)
                if total == 0:
                    continue
                #total = total - 1
                print(attributes)
                print(values)
                if type(values[0]) == str:
                    values[0]=values[0].replace(" ","")
                # have to create the part1 string for the query    
                result=conn.query("MATCH (n:"+ attributes[0] +") WHERE n."+ attributes[0] +"= '"+ str(values[0]) +"'  RETURN count(n)",  db='footballdb')
                for record in result :
                    print(record["count(n)"])
                    if(record["count(n)"] > 0):
                        print("exit")
                        values = []
                        
                    else:
                        p1 = "CREATE( :"+ attributes[0]

                        # have to create the part2 string for the query(attributes)
                        # {"+ attributes[0] +": '"+ values[0] + "' ,"+  attributes[1] +": '" + values[1] + "' ,"+ attributes[2] +":"+ str(values[2]) +"}
                        # {Name: 'Indipendente' , Country: 'Argentina'}
                        p2 = "{"
                        first_attribute = True
                        for i in range(0,len(values)):
                            if type(values[i]) != 'str':
                                values[i] = str(values[i])
                            values[i]=values[i].replace("'","")
                            str_concat = attributes[i] +": '"+ values[i] +"'"
                            if first_attribute == True:
                                first_attribute = False
                                p2 = p2 + str_concat
                            else:
                                p2 = p2 +"," + str_concat
                        p2 = p2 +"} )"
                        values = []
                        conn.query(p1+p2,  db='footballdb')

    elif wtd_option == "CE":
        print("Correct edges")

        total=1
        relation_name = []
        attributes = []
        values = []
        last_read_e = 0 # for resuming when the file is not readed completely
        # Iterate the loop to read the cell values
        for i in range(0, worksheet.max_row): # loop over the rows
            print(str(int(i* 100 /10000)) + "% of the file"+ path_to_file)

            if (i == 0): # CASE relation name
                for col in worksheet.iter_cols(0, worksheet.max_column):
                    relation_name.append(col[i].value)
            elif (i == 1): # CASE when we have the first row (update attributes)
                for col in worksheet.iter_cols(0, worksheet.max_column):
                    attributes.append(col[i].value)
            else: # CASE for rows other than first (update values)
                if i < last_read_e:
                    continue
                for col in worksheet.iter_cols(0, worksheet.max_column):
                    values.append(col[i].value)
    
                print(relation_name)
                print(attributes)
                print(values)
                
                #match (p:Person) where p.Age="22"
                #match (c:ClubID) where c.ClubID="3" 
                #match r=(p)-[:LIKES]->(c) 
                #return count(r)

                exp1="MATCH (first:"+ str(attributes[0]) +") where first."+ str(attributes[0]) +"= '"+ str(values[0]) +"' "
                exp2="MATCH (second:"+ str(attributes[1]) +") where second."+ str(attributes[1]) +"= '"+ str(values[1]) +"' "
                exp3="MATCH r=(first)-[:"+ relation_name[1] +"]->(second) return count(r)"

                print(exp1+exp2+exp3)
                result=conn.query(exp1+exp2+exp3,  db='footballdb')
                for record in result :
                    print(record["count(r)"])
                    if(record["count(r)"] > 0):
                        print("exit")
                        values = []
                        
                    else:
                        # create the outgoing part
                        p1 = "MATCH (first:"+attributes[0] +") where first."+ attributes[0] +"= '"+ str(values[0]) +"' "
                        p2 = ""
                        relation_att =""
                        # create the link part
                        first_attribute = True
                        for j in range(1,len(values)):
                            #print(type(values[j]))
                            if type(values[j]) != str:
                                values[j] = str(values[j])
                            if (j == 1):
                                # create the ingoing node
                                p2 = " MATCH (second:"+attributes[j] +") where second."+ attributes[j] +"= '"+ values[j] +"'"
                            else:
                                if first_attribute == True:
                                    first_attribute = False
                                    relation_att = relation_att + attributes[j] +":"+ values[j]
                                else:
                                    relation_att = relation_att +","+ attributes[j] +":"+ values[j]
                        p3 = " create (first) - [:"+ relation_name[1] + "{"+ relation_att +"} ] -> (second)"
                        conn.query(p1+p2+p3,  db='footballdb')
                        values = []
            
    
    else:
        print("Wrong input")




        
        
