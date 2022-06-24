import openpyxl
import xlsxwriter

if __name__ == "__main__":

    path_to_file = input("Enter the file's path to load: ")  
    print(path_to_file)
    workbook = openpyxl.load_workbook(path_to_file)
    worksheet = workbook.active

    number_of_rows = input("Enter the number of rows per file: ")
    print(number_of_rows)
    number_of_rows = int(number_of_rows)

    header_rows = input("Enter the number of header rows to copy: ")
    print(header_rows)
    header_rows = int(header_rows)

    # Record all the header rows
    header = []
    for i in range(0, header_rows):
        row = []
        for col in worksheet.iter_cols(0, worksheet.max_column):
            row.append(col[i].value)
        header.append(row)
        row = []
    print(header)
    print(header[0])
    print(header[0][0])


    current_coloumn = 0
    current_row = 0
    create_new_file = False
    current_file_number = 0

    # Empty file returned
    wb = xlsxwriter.Workbook('splited/split_'+ str(current_file_number) +'.xlsx')
    ws = wb.add_worksheet()

    for i in range(header_rows , worksheet.max_row):
        if i == 400: 
            break

        # number of rows reached
        if i%number_of_rows == header_rows :
            create_new_file = True
        
        # create a new file and writes the header
        if create_new_file == True:
            current_file_number = current_file_number + 1
            wb.close()
            wb = xlsxwriter.Workbook('splited/split_'+ str(current_file_number) +'.xlsx')
            ws = wb.add_worksheet()
            current_row=-1
            for rowlist in header:
                current_row += 1
                current_coloumn=0
                for elem in rowlist:
                    print(elem)
                    ws.write(current_row, current_coloumn, elem)
                    current_coloumn += 1
            create_new_file = False
            
       
        # Update number of rows and coloumns
        current_row += 1
        current_coloumn = 0
        
        
        for col in worksheet.iter_cols(0, worksheet.max_column):
            ws.write(current_row, current_coloumn, col[i].value)
            current_coloumn += 1

    # close the last workbook
    wb.close()
        

    #wb = xlsxwriter.Workbook('hello.xlsx')
    #ws = wb.add_worksheet()
    #ws.write('A1', 'Hello..')
    #worksheet.write(row, column, item)
    #wb.close()

    #for i in range(0, worksheet.max_row):
    #for col in worksheet.iter_cols(0, worksheet.max_column):





